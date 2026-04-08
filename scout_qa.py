from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "Arial"
C_HEADER_BG = "1F3864"
C_HEADER_FG = "FFFFFF"
C_CAT_BG    = "2E75B6"
C_CAT_FG    = "FFFFFF"
C_ALT       = "EEF3FA"
C_WHITE     = "FFFFFF"
C_PASS      = "C6EFCE"
C_FAIL      = "FFC7CE"
C_BLOCKED   = "FFEB9C"
C_CRIT      = "FFC7CE"
C_HIGH      = "FFEB9C"
C_MED       = "C6EFCE"
C_LOW       = "DDEBF7"
C_SUBHDR    = "BDD7EE"


def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def hdr(ws, cell, value, fg=C_HEADER_FG, bg=C_HEADER_BG, bold=True, size=10,
        halign="center", valign="center"):
    c = ws[cell]
    c.value = value
    c.font = Font(name=FONT, bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=True)


def write_cell(ws, row, col, value, bold=False, color="000000",
               bg=None, wrap=True, halign="left", size=9, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, bold=bold, color=color, size=size, italic=italic)
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.border = thin_border()
    return c


wb = Workbook()

# ─────────────────────────────────────────────────────────────
# SHEET 1 — Overview / Cover
# ─────────────────────────────────────────────────────────────
ws_cover = wb.active
ws_cover.title = "Overview"
for col, w in [("A",3),("B",28),("C",55),("D",18),("E",18)]:
    ws_cover.column_dimensions[col].width = w

ws_cover.row_dimensions[2].height = 50
ws_cover.merge_cells("B2:E2")
c = ws_cover["B2"]
c.value = "Scout Director — QA Test Plan"
c.font = Font(name=FONT, bold=True, size=22, color=C_HEADER_FG)
c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
c.alignment = Alignment(horizontal="center", vertical="center")

meta = [
    ("Product",        "Scout Director v1.0"),
    ("QA Type",        "Real-User Functional Testing"),
    ("Tested By",      ""),
    ("Date",           ""),
    ("Environment",    "Windows 11 · Ollama local · Qdrant 192.168.1.44:6333"),
    ("Entry Criteria", "Ollama running · Qdrant reachable · 1 Godot project available"),
    ("Exit Criteria",  "All P0/P1 tests PASS · No open Critical/High bugs"),
]
row = 4
for k, v in meta:
    ws_cover.row_dimensions[row].height = 20
    write_cell(ws_cover, row, 2, k, bold=True, bg=C_SUBHDR, halign="right")
    ws_cover.merge_cells(f"C{row}:E{row}")
    write_cell(ws_cover, row, 3, v)
    row += 1

row += 1
ws_cover.row_dimensions[row].height = 22
ws_cover.merge_cells(f"B{row}:E{row}")
hdr(ws_cover, f"B{row}", "RESULT LEGEND", bg=C_CAT_BG)
for label, desc, color in [
    ("PASS",    "Test completed, expected result achieved",  C_PASS),
    ("FAIL",    "Expected result NOT achieved — log a bug",  C_FAIL),
    ("BLOCKED", "Cannot execute — dependency/env issue",     C_BLOCKED),
    ("N/A",     "Not applicable for current config",         "DDDDDD"),
]:
    row += 1
    ws_cover.row_dimensions[row].height = 18
    write_cell(ws_cover, row, 2, label, bold=True, bg=color, halign="center")
    ws_cover.merge_cells(f"C{row}:E{row}")
    write_cell(ws_cover, row, 3, desc)

row += 2
ws_cover.row_dimensions[row].height = 22
ws_cover.merge_cells(f"B{row}:E{row}")
hdr(ws_cover, f"B{row}", "BUG SEVERITY GUIDE", bg=C_CAT_BG)
for label, desc, color in [
    ("Critical (P0)", "App crash / data loss / core feature broken",      C_CRIT),
    ("High (P1)",     "Core feature broken, no workaround",               C_HIGH),
    ("Medium (P2)",   "Feature degraded, workaround exists",              C_MED),
    ("Low (P3)",      "Minor UI/UX cosmetic issue",                       C_LOW),
]:
    row += 1
    ws_cover.row_dimensions[row].height = 18
    write_cell(ws_cover, row, 2, label, bold=True, bg=color, halign="center")
    ws_cover.merge_cells(f"C{row}:E{row}")
    write_cell(ws_cover, row, 3, desc)

row += 2
ws_cover.row_dimensions[row].height = 22
ws_cover.merge_cells(f"B{row}:E{row}")
hdr(ws_cover, f"B{row}", "TEST SHEETS", bg=C_CAT_BG)
for sheet_name, desc in [
    ("01 Setup & Launch",       "First launch, services check, project selection"),
    ("02 Indexing",             "Index a project, re-index, partial index"),
    ("03 DB Reset Flow",        "Nuke & Reset → re-index → query pipeline"),
    ("04 Chat & Query",         "Prompt quality, context depth, personas"),
    ("05 Verification Panel",   "Grounding score, hallucination report, efficiency"),
    ("06 Error Handling",       "Offline services, bad collections, edge cases"),
    ("07 Roadmap",              "Roadmap tab create/edit/complete milestones"),
    ("08 Cloud Export",         "Cloud prompt export, XML format, persona variants"),
    ("Bug Log",                 "Log all failures here with severity + reproduction steps"),
    ("Run Tracker",             "Summary dashboard — pass/fail counts per suite"),
]:
    row += 1
    ws_cover.row_dimensions[row].height = 18
    write_cell(ws_cover, row, 2, sheet_name, bold=True, bg=C_ALT)
    ws_cover.merge_cells(f"C{row}:E{row}")
    write_cell(ws_cover, row, 3, desc)


# ─────────────────────────────────────────────────────────────
# Helper — build a standard test sheet
# ─────────────────────────────────────────────────────────────
COLS = [
    ("A", "ID",               5),
    ("B", "Category",         18),
    ("C", "Test Scenario",    32),
    ("D", "Prerequisites",    28),
    ("E", "Steps to Perform", 52),
    ("F", "Expected Result",  42),
    ("G", "Actual Result",    38),
    ("H", "Result",           10),
    ("I", "Severity if Fail", 14),
    ("J", "Notes / Bug Ref",  32),
]
SEV_COLORS = {"Critical": C_CRIT, "High": C_HIGH,
              "Medium": C_MED,    "Low": C_LOW, "": C_WHITE}


def build_test_sheet(wb, title, tests):
    ws = wb.create_sheet(title)
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 28
    for col_letter, col_name, col_width in COLS:
        ws.column_dimensions[col_letter].width = col_width
        hdr(ws, f"{col_letter}1", col_name, size=9)

    row = 2
    counter = 1
    alt = False
    for t in tests:
        if isinstance(t, str) and t.startswith("##"):
            ws.row_dimensions[row].height = 20
            ws.merge_cells(f"A{row}:J{row}")
            c = ws.cell(row=row, column=1, value=t[2:].strip())
            c.font = Font(name=FONT, bold=True, size=10, color=C_CAT_FG)
            c.fill = PatternFill("solid", fgColor=C_CAT_BG)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            row += 1
            alt = False
            continue

        ws.row_dimensions[row].height = 55
        bg = C_ALT if alt else C_WHITE
        alt = not alt
        sev = t.get("severity", "")
        write_cell(ws, row, 1,  f"TC-{counter:03d}", halign="center", bg=bg)
        write_cell(ws, row, 2,  t.get("category", ""), bg=bg)
        write_cell(ws, row, 3,  t.get("scenario", ""), bold=True, bg=bg)
        write_cell(ws, row, 4,  t.get("prereqs", ""),  bg=bg, italic=True, color="555555")
        write_cell(ws, row, 5,  t.get("steps", ""),    bg=bg)
        write_cell(ws, row, 6,  t.get("expected", ""), bg=bg)
        write_cell(ws, row, 7,  "", bg=C_WHITE)
        write_cell(ws, row, 8,  "", bg=C_WHITE, halign="center")
        write_cell(ws, row, 9,  sev, bg=SEV_COLORS.get(sev, C_WHITE), halign="center")
        write_cell(ws, row, 10, "", bg=C_WHITE)
        row += 1
        counter += 1
    return ws


# ─────────────────────────────────────────────────────────────
# SHEET 2 — Setup & Launch
# ─────────────────────────────────────────────────────────────
build_test_sheet(wb, "01 Setup & Launch", [
    "## PRE-LAUNCH CHECKS",
    {"category": "Environment", "scenario": "Ollama service running",
     "prereqs": "Ollama installed",
     "steps": "1. Open terminal\n2. Run: ollama serve\n3. Confirm 'listening on :11434'",
     "expected": "Ollama starts with no errors. Port 11434 accessible.", "severity": "Critical"},
    {"category": "Environment", "scenario": "Qdrant reachable at 192.168.1.44:6333",
     "prereqs": "Docker + Qdrant running on server",
     "steps": "1. Open browser: http://192.168.1.44:6333/dashboard\n2. Dashboard loads\n3. Note collection count",
     "expected": "Qdrant dashboard loads. Collections list shown.", "severity": "Critical"},
    {"category": "Environment", "scenario": "nomic-embed-text model available",
     "prereqs": "Ollama running",
     "steps": "1. Run: ollama list\n2. Confirm 'nomic-embed-text' in list",
     "expected": "Model listed. If missing: ollama pull nomic-embed-text", "severity": "Critical"},
    "## FIRST LAUNCH",
    {"category": "Launch", "scenario": "App launches without Python errors",
     "prereqs": "Ollama + Qdrant running",
     "steps": "1. Run: streamlit run app.py\n2. Browser opens localhost:8501\n3. Check terminal for tracebacks\n4. Check browser console (F12) for JS errors",
     "expected": "App loads. No red error banners. No Python tracebacks.", "severity": "Critical"},
    {"category": "Launch", "scenario": "Sidebar shows Qdrant connected",
     "prereqs": "App running",
     "steps": "1. Look at left sidebar\n2. Check Qdrant host/port display\n3. Confirm collection dropdown populated",
     "expected": "Sidebar shows Qdrant connected. Collection selector shows existing collections.", "severity": "High"},
    {"category": "Launch", "scenario": "Last project auto-loads on restart",
     "prereqs": "Project was previously loaded",
     "steps": "1. Note active project name\n2. Close browser tab\n3. Stop Streamlit (Ctrl+C)\n4. Re-run: streamlit run app.py\n5. Check if project reloads",
     "expected": "Previous project reloads. Chat history preserved.", "severity": "Medium"},
    "## PROJECT SELECTION",
    {"category": "Project", "scenario": "Select a Godot project folder",
     "prereqs": "A Godot project exists on disk",
     "steps": "1. In sidebar use folder picker or path input\n2. Navigate to Godot project\n3. Confirm project name appears in header",
     "expected": "Project name shown. Active collection set. Tabs usable.", "severity": "High"},
    {"category": "Project", "scenario": "Invalid folder path shows error (not crash)",
     "prereqs": "App running",
     "steps": "1. Enter path: C:/does/not/exist\n2. Confirm app response",
     "expected": "Error message shown. App does not crash. Previous state preserved.", "severity": "Medium"},
    {"category": "Project", "scenario": "Collection selector shows all Qdrant collections",
     "prereqs": "Qdrant has 2+ collections",
     "steps": "1. Open Vector DB tab\n2. Check collection dropdown\n3. All Qdrant collections appear",
     "expected": "All collections listed. Dim/distance shown (NOT N/A).", "severity": "High"},
])

# ─────────────────────────────────────────────────────────────
# SHEET 3 — Indexing
# ─────────────────────────────────────────────────────────────
build_test_sheet(wb, "02 Indexing", [
    "## COLLECTION SETUP",
    {"category": "Collection", "scenario": "Initialize new collection shows correct schema",
     "prereqs": "Vector DB tab · Qdrant running",
     "steps": "1. Go to Vector DB tab\n2. Click 'Initialize Collection'\n3. Name: 'test_qa_col'\n4. Select: 768 (nomic-embed-text)\n5. Click Initialize\n6. Select the new collection from dropdown",
     "expected": "Semantic Dimension: 768. Distance Metric: Cosine. NOT N/A. Total Points: 0.", "severity": "Critical"},
    "## INDEXING",
    {"category": "Indexer", "scenario": "Index Godot project end-to-end",
     "prereqs": "Project selected · Collection exists · Ollama running",
     "steps": "1. Go to Indexer tab\n2. Select: nomic-embed-text\n3. Batch: 256, Threads: 4\n4. Click 'Run Indexer'\n5. Watch terminal\n6. Wait for 'Indexing Complete'",
     "expected": "Terminal shows parsed files + batch uploads. 'Indexing Complete' shown. Vector count > 0.", "severity": "Critical"},
    {"category": "Indexer", "scenario": "Vector count increases after indexing",
     "prereqs": "Indexing completed",
     "steps": "1. Go to Vector DB tab\n2. Select indexed collection\n3. Read Total Points",
     "expected": "Total Points > 0. Roughly matches number of functions parsed.", "severity": "High"},
    {"category": "Indexer", "scenario": "Re-index is idempotent (no duplicates)",
     "prereqs": "Project already indexed",
     "steps": "1. Note current vector count\n2. Click 'Run Indexer' again (same project + collection)\n3. Wait for completion\n4. Check vector count",
     "expected": "Vector count unchanged (upsert = no duplicates).", "severity": "High"},
    {"category": "Indexer", "scenario": "Dimension mismatch warning blocks indexing",
     "prereqs": "Collection is 768d",
     "steps": "1. Switch model to mxbai-embed-large (1024d)\n2. Try to click Run Indexer",
     "expected": "Red warning: dimension mismatch. Run Indexer blocked until user resets.", "severity": "High"},
    {"category": "Indexer", "scenario": "Graph DB populated after indexing",
     "prereqs": "Indexing completed",
     "steps": "1. Go to Graph Map tab\n2. Check that Script, Function, Signal nodes appear",
     "expected": "Graph nodes visible. Topic groupings shown. System map renders.", "severity": "High"},
    {"category": "Indexer", "scenario": "BM25 vocab file saved after indexing",
     "prereqs": "Indexing completed",
     "steps": "1. Navigate to: {project_path}/.scout/\n2. Check bm25_vocab.json exists\n3. Open it — confirm 'idf' key with tokens",
     "expected": "bm25_vocab.json exists with IDF vocabulary. File > 1KB.", "severity": "Medium"},
])

# ─────────────────────────────────────────────────────────────
# SHEET 4 — DB Reset Flow
# ─────────────────────────────────────────────────────────────
build_test_sheet(wb, "03 DB Reset Flow", [
    "## NUKE & RESET",
    {"category": "DB Reset", "scenario": "Nuke & Reset wipes both graph and vector DB",
     "prereqs": "Project indexed (vectors + graph exist)",
     "steps": "1. Go to Indexer tab\n2. Click 'Nuke & Reset'\n3. Confirm dialog\n4. Wait for completion",
     "expected": "Both graph and vector DB wiped. Vector count = 0. Graph empty.", "severity": "Critical"},
    {"category": "DB Reset", "scenario": "Collection schema preserved after reset (NOT N/A)",
     "prereqs": "Nuke & Reset just completed",
     "steps": "1. Go to Vector DB tab\n2. Select collection\n3. Check Semantic Dimension and Distance Metric",
     "expected": "Dimension: 768. Distance: Cosine. NEITHER shows N/A.", "severity": "Critical"},
    {"category": "DB Reset", "scenario": "Re-index after reset succeeds",
     "prereqs": "Collection empty after reset",
     "steps": "1. Go to Indexer tab\n2. Click 'Run Indexer'\n3. Wait for completion\n4. Check vector count",
     "expected": "Indexing succeeds. Vector count > 0. No errors.", "severity": "Critical"},
    {"category": "DB Reset", "scenario": "Chat returns real context after reset + re-index",
     "prereqs": "Re-indexed after reset",
     "steps": "1. Go to Chat tab\n2. Type: 'Show me the main player script'\n3. Send\n4. Open 'Cloud AI Prompt' expander\n5. Check context budget line",
     "expected": "Context budget > 0 tokens. Response contains actual code from project.", "severity": "Critical"},
    {"category": "DB Reset", "scenario": "No pipeline errors after clean reset + re-index",
     "prereqs": "Chat response received",
     "steps": "1. Check below the response for ⚠️ pipeline warning expander",
     "expected": "No warning expander. Clean pipeline run.", "severity": "High"},
    "## FIX BUTTONS",
    {"category": "DB Reset", "scenario": "'Fix: Nuke & Reset' from mismatch banner",
     "prereqs": "Dimension mismatch exists (wrong model selected)",
     "steps": "1. Switch to mismatched model\n2. Mismatch banner appears\n3. Click 'Fix: Nuke & Reset'\n4. Check collection",
     "expected": "Collection recreated with correct dimension. Mismatch warning disappears.", "severity": "High"},
    {"category": "DB Reset", "scenario": "Wipe Graph only keeps vectors intact",
     "prereqs": "Project indexed",
     "steps": "1. In Indexer tab click 'Wipe Graph'\n2. Go to Vector DB tab\n3. Check vector count",
     "expected": "Graph empty. Vectors unchanged. Chat still returns vector results.", "severity": "Medium"},
])

# ─────────────────────────────────────────────────────────────
# SHEET 5 — Chat & Query
# ─────────────────────────────────────────────────────────────
build_test_sheet(wb, "04 Chat & Query", [
    "## BASIC QUERY",
    {"category": "Chat", "scenario": "Query returns non-empty context (critical path)",
     "prereqs": "Project indexed · Chat tab open",
     "steps": "1. Type: 'How does the player move?'\n2. Send\n3. Open 'Cloud AI Prompt' expander\n4. Check '[Context budget: ~X tokens used of 24,000]'",
     "expected": "Context budget shows > 0 tokens used. Response mentions real scripts/functions.", "severity": "Critical"},
    {"category": "Chat", "scenario": "Response cites real file names from project",
     "prereqs": "Project indexed",
     "steps": "1. Ask: 'What scripts handle combat?'\n2. Read response\n3. Verify each .gd file mentioned exists on disk",
     "expected": "All file names cited exist in the project. No invented paths.", "severity": "High"},
    {"category": "Chat", "scenario": "Context Depth slider increases retrieved chunks",
     "prereqs": "Project indexed",
     "steps": "1. Set Context Depth to 10 · Ask: 'List all scripts'\n2. Note token count in budget line\n3. Set Context Depth to 80 · Same question\n4. Compare token counts",
     "expected": "Higher depth → more chunks → higher token count in budget line.", "severity": "Medium"},
    {"category": "Chat", "scenario": "Detail Threshold controls FULL vs SYM chunks",
     "prereqs": "Project indexed",
     "steps": "1. Set Detail Threshold to 5 · Send query\n2. Open Cloud AI Prompt\n3. Count [FULL] vs [SYM] tagged chunks",
     "expected": "Only top 5 chunks are [FULL]. Rest are [SYM] (SymCode compressed).", "severity": "Medium"},
    "## PERSONAS",
    {"category": "Persona", "scenario": "Bug Hunter persona uses bug-focus format",
     "prereqs": "Project indexed",
     "steps": "1. Set AI Archetype: Bug Hunter\n2. Ask: 'Find issues in the player code'\n3. Check response structure",
     "expected": "Response uses: file | function | root cause | fix format. Severity levels listed.", "severity": "Medium"},
    {"category": "Persona", "scenario": "Senior Architect gives structural advice",
     "prereqs": "Project indexed",
     "steps": "1. Set AI Archetype: Senior Architect\n2. Ask: 'Review the overall architecture'",
     "expected": "Response discusses coupling, signal buses, Godot patterns. Not a bug list.", "severity": "Medium"},
    {"category": "Persona", "scenario": "Teacher explains concepts with code examples",
     "prereqs": "Project indexed",
     "steps": "1. Set AI Archetype: Teacher\n2. Ask: 'Explain how signals work in this project'",
     "expected": "Response explains concepts with quotes from actual code. Beginner-friendly tone.", "severity": "Low"},
    "## MULTI-TURN",
    {"category": "Chat", "scenario": "Follow-up question references prior context",
     "prereqs": "1 message already sent",
     "steps": "1. Ask: 'How does the player take damage?'\n2. After response, ask: 'Show me the full function body'",
     "expected": "Follow-up references the take_damage function from prior turn. No re-introduction.", "severity": "Medium"},
    {"category": "Chat", "scenario": "Chat history persists after page refresh",
     "prereqs": "2+ messages in chat",
     "steps": "1. Note visible messages\n2. Refresh browser (F5)\n3. Check messages reload",
     "expected": "Chat history reloads. Messages visible. Project state preserved.", "severity": "High"},
    {"category": "Chat", "scenario": "System map appears for multi-system queries",
     "prereqs": "Project has Topics/signals in graph",
     "steps": "1. Ask: 'How does UI update when player health changes?'\n2. Open Cloud AI Prompt expander\n3. Look for [SYSTEM MAP] section",
     "expected": "[SYSTEM MAP] section shows subsystems, scripts, and shared signals.", "severity": "Medium"},
])

# ─────────────────────────────────────────────────────────────
# SHEET 6 — Verification Panel
# ─────────────────────────────────────────────────────────────
build_test_sheet(wb, "05 Verification Panel", [
    "## GROUNDING SCORE",
    {"category": "Verification", "scenario": "Grounding score caption appears after every response",
     "prereqs": "Project indexed · response received",
     "steps": "1. Send any query\n2. Wait for response\n3. Look immediately below response text",
     "expected": "Caption shows e.g. '✅ Grounding: 85% — 12 verified, 2 unverified'. Icon color matches threshold.", "severity": "High"},
    {"category": "Verification", "scenario": "Score icon colors reflect threshold",
     "prereqs": "Multiple queries sent",
     "steps": "1. Note score icon color across different queries\n2. Verify: ≥90% = ✅  70–89% = ⚠️  <70% = 🔴",
     "expected": "Icons match score ranges consistently across queries.", "severity": "Medium"},
    "## VERIFICATION REPORT",
    {"category": "Verification", "scenario": "Verification report expander opens correctly",
     "prereqs": "Response received",
     "steps": "1. Click '📊 Verification Report' expander",
     "expected": "Expander opens. Shows: Grounding Score, Context Coverage, Entity Precision table, Potential Hallucinations.", "severity": "High"},
    {"category": "Verification", "scenario": "Entity precision breakdown shows per-type rows",
     "prereqs": "Verification report open",
     "steps": "1. Open Verification Report\n2. Look at Entity Precision section",
     "expected": "Rows for: function, signal, file, variable, class. Each row: verified count, hallucinated count, precision %.", "severity": "Medium"},
    {"category": "Verification", "scenario": "Hallucination list flags non-existent references",
     "prereqs": "LLM mentioned something not in graph",
     "steps": "1. Ask about an obscure feature the LLM might invent\n2. Open Verification Report\n3. Check Potential Hallucinations section",
     "expected": "Section lists entities mentioned by LLM but NOT found in knowledge graph.", "severity": "High"},
    {"category": "Verification", "scenario": "Context coverage fraction is accurate",
     "prereqs": "Response received",
     "steps": "1. Open Verification Report\n2. Note Context Coverage (e.g. 3/5 files)\n3. Cross-check: count distinct files in Cloud AI Prompt expander",
     "expected": "Coverage = files cited by LLM ÷ files in prompt context. Fraction matches.", "severity": "Medium"},
    "## PROMPT EFFICIENCY",
    {"category": "Efficiency", "scenario": "Prompt efficiency metrics all non-zero",
     "prereqs": "Verification report open",
     "steps": "1. Scroll to 'Prompt Efficiency' section\n2. Check all 5 metrics",
     "expected": "Budget utilization %, Response density, Unique entities cited, Prompt tokens est, Response tokens est — all numeric, none 0 or N/A.", "severity": "Medium"},
    {"category": "Efficiency", "scenario": "Budget utilization scales with Context Depth",
     "prereqs": "Project indexed",
     "steps": "1. Context Depth=10 · send query · note budget utilization\n2. Context Depth=80 · same query · compare",
     "expected": "Higher depth → higher budget utilization %.", "severity": "Low"},
    "## PIPELINE LOG",
    {"category": "Pipeline Log", "scenario": "No warning shown on clean indexed project",
     "prereqs": "Project fully indexed · Qdrant healthy",
     "steps": "1. Send a query\n2. Check if ⚠️ pipeline warning expander appears",
     "expected": "No warning expander visible. Clean run.", "severity": "Medium"},
    {"category": "Pipeline Log", "scenario": "Pipeline warning shown when graph is empty",
     "prereqs": "Graph wiped, vectors remain",
     "steps": "1. Wipe graph only\n2. Send a query\n3. Check for ⚠️ expander",
     "expected": "Warning expander appears. Shows graph query failures. Response still returned.", "severity": "Medium"},
])

# ─────────────────────────────────────────────────────────────
# SHEET 7 — Error Handling
# ─────────────────────────────────────────────────────────────
build_test_sheet(wb, "06 Error Handling", [
    "## QDRANT OFFLINE",
    {"category": "Error Handling", "scenario": "Query handles Qdrant offline gracefully",
     "prereqs": "App running with project loaded",
     "steps": "1. Disconnect from 192.168.1.44 / stop Qdrant\n2. Send a query in chat\n3. Check response + pipeline log",
     "expected": "Pipeline error shown: 'Qdrant search failed'. Response still generated (graph-only). No crash.", "severity": "High"},
    {"category": "Error Handling", "scenario": "Indexer shows error when Qdrant offline",
     "prereqs": "Qdrant offline",
     "steps": "1. Try to run indexer\n2. Note error display",
     "expected": "Clear error message (not Python traceback). App doesn't freeze.", "severity": "High"},
    "## OLLAMA OFFLINE",
    {"category": "Error Handling", "scenario": "Chat shows clear error when Ollama offline",
     "prereqs": "Ollama stopped",
     "steps": "1. Stop: ollama serve\n2. Send a query\n3. Read response",
     "expected": "Response: '❌ Local Engine Error: ...' with instructions. Not a crash.", "severity": "High"},
    {"category": "Error Handling", "scenario": "Embedding fails gracefully when Ollama offline",
     "prereqs": "Ollama stopped",
     "steps": "1. Try to run indexer with Ollama offline\n2. Check UI feedback",
     "expected": "Error shown. Indexer stops cleanly. No partial/corrupted data written.", "severity": "High"},
    "## BAD COLLECTION STATE",
    {"category": "Error Handling", "scenario": "Legacy unnamed vector collection shows warning",
     "prereqs": "Collection with unnamed vectors exists (e.g. floodedcafe11)",
     "steps": "1. Select floodedcafe11 as active collection\n2. Send a query\n3. Check pipeline log",
     "expected": "Warning: 'Collection uses legacy unnamed vectors. Re-index with Nuke & Reset.' Response still attempted.", "severity": "Medium"},
    {"category": "Error Handling", "scenario": "Empty collection returns graph-only warning",
     "prereqs": "Collection with 0 vectors selected",
     "steps": "1. Select an empty collection\n2. Send query",
     "expected": "Warning: 'No vector results returned — context will be graph-only.' No crash.", "severity": "Medium"},
    "## EDGE CASES",
    {"category": "Edge Case", "scenario": "Very long query (2000+ chars) doesn't crash",
     "prereqs": "Project indexed",
     "steps": "1. Paste 2000+ character query\n2. Send",
     "expected": "App sends query. Response received. No timeout or Python exception.", "severity": "Medium"},
    {"category": "Edge Case", "scenario": "Empty query does not trigger API call",
     "prereqs": "Chat tab open",
     "steps": "1. Click send without typing\n2. Or press Enter on empty input",
     "expected": "Nothing sent. No spinner. Input remains empty.", "severity": "Low"},
    {"category": "Edge Case", "scenario": "Project with no .gd files indexes without crash",
     "prereqs": "Folder with only .md/.json files",
     "steps": "1. Select folder with no .gd files\n2. Run indexer",
     "expected": "Indexer completes. Low/zero vector count. No crash. Markdown/JSON parsed.", "severity": "Medium"},
    {"category": "Edge Case", "scenario": "Context Window setting persists after refresh",
     "prereqs": "App running",
     "steps": "1. Set Context Window slider to 16384\n2. Refresh page\n3. Check slider value",
     "expected": "Slider retains 16384. Preference saved to user_prefs.json.", "severity": "Low"},
])

# ─────────────────────────────────────────────────────────────
# SHEET 8 — Roadmap
# ─────────────────────────────────────────────────────────────
build_test_sheet(wb, "07 Roadmap", [
    "## ROADMAP TAB",
    {"category": "Roadmap", "scenario": "Roadmap tab renders without error",
     "prereqs": "Project selected",
     "steps": "1. Click Roadmap tab\n2. Check it renders",
     "expected": "Tab renders. Global goal field visible. Milestones section present.", "severity": "Medium"},
    {"category": "Roadmap", "scenario": "Set global project goal persists",
     "prereqs": "Roadmap tab open",
     "steps": "1. Enter: 'Build a 2D platformer'\n2. Save\n3. Refresh page",
     "expected": "Goal reappears after refresh. Shows in chat prompt under [PROJECT ROADMAP].", "severity": "Medium"},
    {"category": "Roadmap", "scenario": "Add and save a new milestone",
     "prereqs": "Roadmap tab open",
     "steps": "1. Enter milestone: 'Player Movement'\n2. Status: Pending\n3. Save\n4. Refresh",
     "expected": "Milestone appears in list. Status shown. Persists after refresh.", "severity": "Medium"},
    {"category": "Roadmap", "scenario": "Completed milestone appears in chat prompt",
     "prereqs": "1 milestone exists",
     "steps": "1. Change milestone status to Completed\n2. Send a chat query\n3. Open Cloud AI Prompt expander",
     "expected": "Prompt contains milestone name under 'Completed:' in [PROJECT ROADMAP] section.", "severity": "Low"},
    {"category": "Roadmap", "scenario": "Delete milestone removes it permanently",
     "prereqs": "1+ milestones exist",
     "steps": "1. Delete a milestone\n2. Refresh page\n3. Confirm it's gone",
     "expected": "Milestone removed. Does not reappear after refresh.", "severity": "Low"},
])

# ─────────────────────────────────────────────────────────────
# SHEET 9 — Cloud Export
# ─────────────────────────────────────────────────────────────
build_test_sheet(wb, "08 Cloud Export", [
    "## CLOUD EXPORT",
    {"category": "Cloud Export", "scenario": "Cloud AI Prompt expander shows XML-tagged prompt",
     "prereqs": "Query sent",
     "steps": "1. Click 'Cloud AI Prompt' expander\n2. Scroll through content",
     "expected": "XML sections visible: <ARCHITECTURAL_CONTEXT>, <USER_QUERY>, <INSTRUCTIONS>.", "severity": "Medium"},
    {"category": "Cloud Export", "scenario": "Cloud Export tab generates full codebase prompt",
     "prereqs": "Project selected with .gd files",
     "steps": "1. Go to Cloud Export tab\n2. Enter query: 'Full architecture review'\n3. Click Export/Generate",
     "expected": "Full XML prompt with all project files. <file path=...> tags for each file.", "severity": "Medium"},
    {"category": "Cloud Export", "scenario": "Copy to clipboard works without truncation",
     "prereqs": "Cloud export generated",
     "steps": "1. Click 'Copy to Clipboard'\n2. Paste into text editor\n3. Verify content",
     "expected": "Full prompt pasted. No truncation. Valid XML.", "severity": "Low"},
    {"category": "Cloud Export", "scenario": "Bug Hunter persona changes cloud prompt instructions",
     "prereqs": "Cloud Export tab",
     "steps": "1. Generate with Senior Architect\n2. Copy <INSTRUCTIONS> section\n3. Switch to Bug Hunter\n4. Generate again\n5. Compare <INSTRUCTIONS>",
     "expected": "<INSTRUCTIONS> differ: architecture critique vs bug hunting format.", "severity": "Low"},
])

# ─────────────────────────────────────────────────────────────
# SHEET 10 — Bug Log
# ─────────────────────────────────────────────────────────────
ws_bug = wb.create_sheet("Bug Log")
ws_bug.freeze_panes = "A3"

bug_cols = [
    ("A", "Bug ID",              8),
    ("B", "Date Found",          12),
    ("C", "Found By",            14),
    ("D", "Test Case",           10),
    ("E", "Module / Tab",        16),
    ("F", "Severity",            12),
    ("G", "Title",               40),
    ("H", "Steps to Reproduce",  50),
    ("I", "Expected",            30),
    ("J", "Actual",              30),
    ("K", "Screenshot",          16),
    ("L", "Status",              12),
    ("M", "Fixed In",            12),
    ("N", "Notes",               30),
]
ws_bug.row_dimensions[1].height = 28
for col_letter, col_name, col_width in bug_cols:
    ws_bug.column_dimensions[col_letter].width = col_width
    hdr(ws_bug, f"{col_letter}1", col_name, size=9)

known_bugs = [
    ("2026-04-06","Dev","TC-001","Vector DB Tab","Critical",
     "Dimension / Distance shows N/A after collection creation",
     "1. Vector DB tab\n2. Initialize Collection (default settings)\n3. Select it",
     "768 / Cosine", "N/A / N/A", "FIXED", "Hotfix 2026-04-06",
     "Display code checked 'semantic' key; collection uses 'dense'"),
    ("2026-04-06","Dev","TC-002","Chat Tab","Critical",
     "Vector search 400 Bad Request after UI-created collection",
     "1. Create collection via Initialize button\n2. Index project\n3. Send any query",
     "Results returned", "400: Not existing vector name",
     "FIXED", "Hotfix 2026-04-06",
     "UI created unnamed VectorParams; fixed to named dense+bm25 schema"),
    ("2026-04-06","Dev","TC-003","Sidebar","Medium",
     "StreamlitAPIException: ui_num_ctx widget key conflict",
     "1. Load app\n2. Check terminal for Streamlit warnings",
     "No warnings", "StreamlitAPIException on ui_num_ctx",
     "FIXED", "Hotfix 2026-04-06",
     "Removed value= from select_slider; guard with 'if key not in session_state'"),
    ("2026-04-06","Dev","TC-004","Graph DB","High",
     "Kuzu RuntimeError: Cannot create empty DB under READ ONLY mode",
     "1. Fresh scout_dir (no graph)\n2. Send a query",
     "Graph opened empty, queries return nothing",
     "RuntimeError crash",
     "FIXED", "Hotfix 2026-04-06",
     "graph_db.py now checks db_exists before setting effective_read_only"),
    ("2026-04-06","Dev","TC-005","Ask.py","Critical",
     "QdrantClient.search() removed in v1.17 — legacy fallback crashes",
     "1. Upgrade qdrant-client to v1.17\n2. Query legacy unnamed collection",
     "Fallback returns results",
     "AttributeError: QdrantClient has no .search()",
     "FIXED", "Hotfix 2026-04-06",
     "Replaced .search() with .query_points(query=vec, no using=)"),
]

sev_map = {"Critical": C_CRIT, "High": C_HIGH, "Medium": C_MED, "Low": C_LOW}
status_map = {"FIXED": C_PASS, "OPEN": C_FAIL, "IN PROGRESS": C_BLOCKED}

for i, bug in enumerate(known_bugs, start=1):
    row = i + 1
    ws_bug.row_dimensions[row].height = 60
    date, found_by, tc, module, sev, title, steps, expected, actual, status, fixed_in, notes = bug
    vals = [f"BUG-{i:03d}", date, found_by, tc, module, sev, title, steps,
            expected, actual, "", status, fixed_in, notes]
    for col, val in enumerate(vals, start=1):
        bg = C_WHITE
        if col == 6:  bg = sev_map.get(sev, C_WHITE)
        if col == 12: bg = status_map.get(status, C_WHITE)
        write_cell(ws_bug, row, col, val, bg=bg, size=8)

for i in range(len(known_bugs)+1, len(known_bugs)+21):
    row = i + 1
    ws_bug.row_dimensions[row].height = 40
    write_cell(ws_bug, row, 1, f"BUG-{i:03d}", halign="center", bg=C_ALT, size=8)
    for col in range(2, 15):
        write_cell(ws_bug, row, col, "", bg=C_WHITE, size=8)

# ─────────────────────────────────────────────────────────────
# SHEET 11 — Run Tracker
# ─────────────────────────────────────────────────────────────
ws_track = wb.create_sheet("Run Tracker")
for col, w in [("A",3),("B",30),("C",14),("D",14),("E",14),("F",14),("G",14)]:
    ws_track.column_dimensions[col].width = w

ws_track.row_dimensions[1].height = 40
ws_track.merge_cells("B1:G1")
c = ws_track["B1"]
c.value = "Scout Director — QA Run Tracker"
c.font = Font(name=FONT, bold=True, size=18, color=C_HEADER_FG)
c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
c.alignment = Alignment(horizontal="center", vertical="center")

# Meta
ws_track.row_dimensions[2].height = 8
for row, (k, v) in enumerate([("Tester",""),("Run Date",""),("Build","v1.0")], start=3):
    ws_track.row_dimensions[row].height = 20
    write_cell(ws_track, row, 2, k, bold=True, bg=C_SUBHDR, halign="right")
    ws_track.merge_cells(f"C{row}:G{row}")
    write_cell(ws_track, row, 3, v)

# Header row
hdr_row = 7
ws_track.row_dimensions[hdr_row].height = 22
for col, label in enumerate(["Test Suite","Total TCs","Pass","Fail","Blocked","Pass Rate"], start=2):
    write_cell(ws_track, hdr_row, col, label, bold=True, bg=C_SUBHDR, halign="center")

suites = [
    ("01 Setup & Launch",     9),
    ("02 Indexing",           7),
    ("03 DB Reset Flow",      7),
    ("04 Chat & Query",      10),
    ("05 Verification Panel",10),
    ("06 Error Handling",    10),
    ("07 Roadmap",            5),
    ("08 Cloud Export",       4),
]

data_start = hdr_row + 1
for i, (suite, count) in enumerate(suites):
    row = data_start + i
    ws_track.row_dimensions[row].height = 20
    bg = C_ALT if i % 2 == 0 else C_WHITE
    write_cell(ws_track, row, 2, suite, bold=True, bg=bg)
    write_cell(ws_track, row, 3, count, halign="center", bg=bg)
    write_cell(ws_track, row, 4, "", halign="center", bg=C_WHITE)
    write_cell(ws_track, row, 5, "", halign="center", bg=C_WHITE)
    write_cell(ws_track, row, 6, "", halign="center", bg=C_WHITE)
    c = ws_track.cell(row=row, column=7,
                      value=f"=IFERROR(D{row}/C{row},\"-\")")
    c.font = Font(name=FONT, size=9)
    c.number_format = "0%"
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=C_WHITE)
    c.border = thin_border()

total_row = data_start + len(suites)
ws_track.row_dimensions[total_row].height = 24
write_cell(ws_track, total_row, 2, "TOTAL", bold=True,
           bg=C_HEADER_BG, color=C_HEADER_FG, halign="center")
for col, formula in [
    (3, f"=SUM(C{data_start}:C{total_row-1})"),
    (4, f"=SUM(D{data_start}:D{total_row-1})"),
    (5, f"=SUM(E{data_start}:E{total_row-1})"),
    (6, f"=SUM(F{data_start}:F{total_row-1})"),
]:
    c = ws_track.cell(row=total_row, column=col, value=formula)
    c.font = Font(name=FONT, bold=True, size=9, color=C_HEADER_FG)
    c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()

c = ws_track.cell(row=total_row, column=7,
                  value=f"=IFERROR(D{total_row}/C{total_row},\"-\")")
c.font = Font(name=FONT, bold=True, size=9, color=C_HEADER_FG)
c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
c.number_format = "0%"
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = thin_border()

# Bug summary
row = total_row + 2
ws_track.row_dimensions[row].height = 22
ws_track.merge_cells(f"B{row}:G{row}")
hdr(ws_track, f"B{row}", "OPEN BUG SUMMARY  (update manually from Bug Log tab)", bg=C_CAT_BG)
for label, color in [("Critical Open", C_CRIT), ("High Open", C_HIGH),
                      ("Medium Open", C_MED),    ("Low Open", C_LOW)]:
    row += 1
    ws_track.row_dimensions[row].height = 18
    write_cell(ws_track, row, 2, label, bold=True, bg=color)
    write_cell(ws_track, row, 3, 0, halign="center", bg=C_WHITE)
    ws_track.merge_cells(f"D{row}:G{row}")
    write_cell(ws_track, row, 4, "Count open bugs of this severity", italic=True, color="888888")

# ─────────────────────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────────────────────
out = "D:/Sk0uter/Scout_Director_QA_TestPlan.xlsx"
wb.save(out)
print(f"Saved: {out}")
